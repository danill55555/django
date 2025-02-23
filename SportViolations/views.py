from django.shortcuts import render, redirect
from .models import Profile
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from .models import Questionnaire
# Create your views here.
def index(request):
    return render(request, 'SportViolations/index.html')

def register(request):
    if request.method == 'POST':
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        patronymic = request.POST['patronymic']
        email = request.POST['email']
        birth_date = request.POST['birth_date']
        password = request.POST['password']

        user = User.objects.create_user(username=email, password=password)
        Profile.objects.create(user=user, first_name=first_name, last_name=last_name, patronymic=patronymic, email=email, birth_date=birth_date)

        messages.success(request, 'Регистрация прошла успешно!')
        return redirect('home')  # На страницу входа после успешной регистрации

    return render(request, 'SportViolations/register.html')

from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages


from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages

def login_view(request):
    # Проверка, аутентифицирован ли пользователь
    if request.user.is_authenticated:
        return redirect('home')  # Перенаправление на главную страницу, если пользователь уже вошел

    if request.method == 'POST':
        username = request.POST['username']  # Запрос имени пользователя
        password = request.POST['password']  # Запрос пароля

        # Автентификация пользователя
        user = authenticate(username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request, 'Вы успешно вошли в систему!')
            return redirect('home')  # Перенаправление на домашнюю страницу
        else:
            messages.error(request, 'Неправильное имя пользователя или пароль.')

    return render(request, 'SportViolations/login.html')


# Выход из аккаунта пользователей
def logout_view(request):
    logout(request)
    return redirect('home')

# Личный кабинет
from django.core.paginator import Paginator

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.contrib import messages
from .models import Questionnaire, Anthropometry
from django.core.paginator import Paginator

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Account, Questionnaire  # Импортируем модель Account
from django.core.paginator import Paginator
from datetime import date

from django.shortcuts import render
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from .models import Questionnaire, Account


from django.shortcuts import render
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from .models import Questionnaire, Account

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.core.paginator import Paginator
from .models import Questionnaire, Account

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.core.paginator import Paginator
from .models import Questionnaire, Account




from django.db.models import Q
from datetime import datetime

from django.shortcuts import render
from django.db.models import Q
from django.contrib.auth.decorators import login_required


from django.shortcuts import render
from django.db.models import Q
from django.contrib.auth.decorators import login_required

from django.shortcuts import render
from django.db.models import Q
from django.contrib.auth.decorators import login_required

from django.shortcuts import render, get_object_or_404
from django.db.models import Q
from django.contrib.auth.decorators import login_required

from django.shortcuts import render
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.http import Http404

@login_required
def profile(request):
    # Получаем фильтры из запроса
    name_filter = request.GET.get('name_filter', '')
    registration_date_filter = request.GET.get('registration_date_filter', '')
    year_filter = request.GET.get('year_filter', '')
    trainer_filter = request.GET.get('trainer_filter', '')
    experience_filter = request.GET.get('experience_filter', '')
    filter_active = request.GET.get('filter_active', 'false') == 'true'

    # Проверка на то, является ли пользователь администратором
    is_admin = request.user.is_staff

    # Получаем объект Account для текущего пользователя
    user_account = Account.objects.filter(user=request.user).first()  # Используем first() для безопасного извлечения

    # Начальная выборка студентов
    if is_admin:
        students = Questionnaire.objects.all()  # Все анкеты для администраторов
    else:
        students = Questionnaire.objects.filter(trainer=user_account) if user_account else Questionnaire.objects.none()  # Только свои анкеты для тренеров, если аккаунт существует

    # Применение фильтров
    if name_filter:
        students = students.filter(Q(name__icontains=name_filter) |
                                   Q(surname__icontains=name_filter) |
                                   Q(patronymic__icontains=name_filter))
    if registration_date_filter:
        students = students.filter(registration_date__date=registration_date_filter)
    if year_filter:
        students = students.filter(registration_date__year=year_filter)
    if experience_filter:
        students = students.filter(experience__icontains=experience_filter)
    if filter_active:
        students = students.filter(is_created=True).distinct()

    # Получение уникальных годов регистрации
    registration_years = sorted({student.registration_date.year for student in students if student.registration_date})

    # Группировка студентов по дате регистрации
    grouped_students = []
    registration_dates = sorted({student.registration_date.date() for student in students})

    # Получаем уникальные даты
    for registration_date in registration_dates:
        students_for_date = students.filter(registration_date__date=registration_date)  # Фильтруем студентов по дате
        grouped_students.append((registration_date, students_for_date))  # Добавляем кортеж (дата, студенты)

    # Определение количества найденных записей
    result_count = students.count()

    # Передаем данные в шаблон и возвращаем HttpResponse
    return render(request, 'SportViolations/profile.html', {
        'students': students,
        'grouped_students': grouped_students,
        'registration_years': registration_years,
        'result_count': result_count,
        'is_admin': is_admin,
    })




# @login_required
# def profile(request):
#     # Получаем фильтры из запроса
#     name_filter = request.GET.get('name_filter', '')
#     date_filter = request.GET.get('date_filter', '')
#     trainer_filter = request.GET.get('trainer_filter', '')
#     experience_filter = request.GET.get('experience_filter', '')
#     filter_active = request.GET.get('filter_active', 'false') == 'true'
#
#     # Проверка на то, является ли пользователь администратором
#     is_admin = request.user.is_staff
#
#     # Начальная выборка студентов
#     students = Questionnaire.objects.all()
#
#     # Применение фильтров
#     if name_filter:
#         students = students.filter(
#             Q(name__icontains=name_filter) |
#             Q(surname__icontains=name_filter) |
#             Q(patronymic__icontains=name_filter)
#         )
#
#     if date_filter:
#         students = students.filter(date_of_birth=date_filter)
#
#     if experience_filter:
#         students = students.filter(experience__icontains=experience_filter)
#
#     if filter_active:
#         students = students.filter(anthropometry__is_created=True).distinct()
#
#     if is_admin and trainer_filter:
#         accounts = Account.objects.filter(
#             Q(surname__icontains=trainer_filter) |
#             Q(user__username__icontains=trainer_filter) |  # Исправлено
#             Q(patronymic__icontains=trainer_filter),
#             role='trainer'
#         )
#         students = students.filter(trainer__in=accounts)
#
#     # Определение количества найденных записей
#     result_count = students.count()
#
#     # Пагинация
#     paginator = Paginator(students, 14)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)
#
#     # Проверка наличия результатов
#     has_results = page_obj.object_list.exists()
#
#     # Получаем информацию о тренерах
#     accounts = Account.objects.filter(role='trainer')
#
#     # Определяем, есть ли анкеты и извлекаем дату регистрации
#     if students.exists():
#         registration_date = students.first().registration_date  # Получаем дату регистрации первой анкеты
#         registration_year = registration_date.year  # Извлекаем год
#         registration_month = registration_date.strftime('%B')  # Извлекаем месяц в текстовом виде (например, "Январь")
#         registration_day = registration_date.day  # Извлекаем день
#         formatted_date = registration_date.strftime('%d.%m.%Y')  # Форматируем дату для вывода
#     else:
#         registration_year = None
#         registration_month = None
#         registration_day = None
#         formatted_date = None
#
#     # Отправляем объекты в шаблон
#     return render(request, 'SportViolations/profile.html', {
#         'page_obj': page_obj,
#         'name_filter': name_filter,
#         'date_filter': date_filter,
#         'trainer_filter': trainer_filter,
#         'experience_filter': experience_filter,
#         'filter_active': filter_active,
#         'is_admin': is_admin,
#         'accounts': accounts,
#         'result_count': result_count,
#         'has_results': has_results,  # Добавлено для проверки наличия результатов
#           'registration_year': registration_year,
#         'registration_month': registration_month,
#         'registration_day': registration_day,
#         'formatted_date': formatted_date
#     })





import openpyxl
from django.http import HttpResponse

import xlwt
from django.http import HttpResponse
from .models import Questionnaire

from django.http import HttpResponse
from openpyxl import Workbook
from .models import Questionnaire


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Questionnaire

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Questionnaire

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.db.models import Q


def export_to_excel(request):
    # Получение параметров фильтрации из запроса
    name_filter = request.GET.get('name_filter', '')
    date_filter = request.GET.get('date_filter', '')
    filter_active = request.GET.get('filter_active', 'false').lower() == 'true'

    # Извлечение всех студентов
    students = Questionnaire.objects.all()

    # Применение фильтра по имени
    if name_filter:
        students = students.filter(
            Q(name__icontains=name_filter) |
            Q(surname__icontains=name_filter) |
            Q(patronymic__icontains=name_filter)
        )

    # Применение фильтра по дате рождения
    if date_filter:
        date_parts = date_filter.split('-')
        year, month, day = int(date_parts[0]), int(date_parts[1]), int(date_parts[2])
        students = students.filter(date_of_birth__year=year, date_of_birth__month=month, date_of_birth__day=day)

    # Фильтр только активных студентов
    if filter_active:
        students = students.filter(anthropometry__is_created=True).distinct()

    # Создаем Excel-файл
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Список спортсменов'

    # Устанавливаем заголовки колонок
    headers = ['ФИО', 'Дата рождения', 'Стаж занятий', 'Тренер', 'Паспортичка', 'Тестирование']
    worksheet.append(headers)

    # Стилизация заголовков
    for cell in worksheet["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")  # Жирный текст белого цвета
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Синий цвет фона
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Центрирование текста
        cell.border = Border(bottom=Side(style='thick', color="000000"))  # Толстая граница снизу

    # Заполняем данные студентов
    for student in students:
        full_name = f"{student.surname} {student.name} {student.patronymic}"

        # Получаем информацию о тренере
        if student.trainer:
            trainer_info = f"{student.trainer.surname} {student.trainer.patronymic} {student.trainer.user.username}"
        else:
            trainer_info = 'N/A'

        # Получаем статусы
        testing_result_symbol = '✔' if student.is_created else '✖'

        # Цвет обозначения
        testing_result_symbol_cell = Font(color="008000") if student.is_created else Font(color="FF0000")

        # Проверка наличия объектов Anthropometry
        if student.anthropometry.exists():
            is_created_status_symbol = '✔' if student.anthropometry.first().is_created else '✖'
            is_created_status_symbol_cell = Font(color="008000") if student.anthropometry.first().is_created else Font(
                color="FF0000")
        else:
            is_created_status_symbol = '✖'
            is_created_status_symbol_cell = Font(color="FF0000")

        # Добавляем строку в Excel
        worksheet.append([
            full_name,
            student.date_of_birth.strftime('%Y-%m-%d'),

            student.qualification,  # Стаж занятий
            trainer_info,  # Тренер
            testing_result_symbol,  # Статус тестирования
            is_created_status_symbol  # Статус антропометрии
        ])

        # Получаем индексы последних добавленных ячеек
        last_row = worksheet.max_row

        # Применяем цвет к статусам
        worksheet.cell(row=last_row, column=6).font = testing_result_symbol_cell  # Тестирование
        worksheet.cell(row=last_row, column=7).font = is_created_status_symbol_cell  # Антропометрия

    # Автоматическая ширина колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Получаем букву колонки
        for cell in column:
            try:
                # Проверяем длину текста в ячейке и обновляем максимальную ширину
                if cell.value:  # Проверяем только если ячейка не пустая
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Устанавливаем ширину колонки, добавляя немного пространства
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Создание HTTP Response для отправки файла Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sport.xlsx'  # Имя файла

    # Сохраняем рабочую книгу в ответ
    workbook.save(response)
    return response


from .forms import StudentForm
from django.http import HttpResponseRedirect
from django.urls import reverse

from django.core.files.storage import default_storage


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required

@login_required
def student_form(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            student = form.save(commit=False)
            student.trainer = Account.objects.get(user=request.user)  # Сохраняем тренера, создавшего анкету
            student.is_created = True  # Устанавливаем статус создания
            student.save()  # Сохраняем в базу данных
            return redirect('assessment', student_id=student.survey_id)  # Перенаправляем на нужную страницу
    else:
        form = StudentForm()

    return render(request, 'SportViolations/questionnaire.html', {'form': form})


from django.shortcuts import render, redirect
from .forms import AnthropometryForm, WristDynamometryForm
from .models import Anthropometry, WristDynamometry

def assessment_view(request, student_id):
    student = Questionnaire.objects.get(survey_id=student_id)

    if request.method == 'POST':
        anthropometry_form = AnthropometryForm(request.POST)
        specific_abilities_form = WristDynamometryForm(request.POST)
        if anthropometry_form.is_valid() and specific_abilities_form.is_valid():
            anthropometry = anthropometry_form.save(commit=False)
            anthropometry.student = student
            anthropometry.is_created = True  # Устанавливаем is_created в True
            anthropometry.save()

            specific_abilities = specific_abilities_form.save(commit=False)
            specific_abilities.student = student
            specific_abilities.save()

            return redirect('result')
    else:
        anthropometry_form = AnthropometryForm()
        specific_abilities_form = WristDynamometryForm()

    return render(request, 'SportViolations/assessment.html', {
        'anthropometry_form': anthropometry_form,
        'specific_abilities_form': specific_abilities_form,
        'student': student
    })


def result_view(request):
    # Здесь можно добавить логику для отображения результатов
    return render(request, 'result.html')

from django.shortcuts import render, get_object_or_404


def personal_account(request, student_id):
    # Получаем объект анкеты студента по его уникальному идентификатору
    student = get_object_or_404(Questionnaire, survey_id=student_id)

    # Получаем связанные данные о динамометрии
    wrist_dynamometry = WristDynamometry.objects.filter(student=student).first()

    # Получаем связанные данные о антропометрии
    anthropometry = Anthropometry.objects.filter(student=student).first()

    # Передаем данные в шаблон для отображения
    return render(request, 'SportViolations/personal_account.html', {
        'student': student,
        'wrist_dynamometry': wrist_dynamometry,
        'anthropometry': anthropometry,
    })


from django.shortcuts import render, redirect
from .forms import TrainerForm
from django.shortcuts import render, redirect
from .forms import TrainerForm
from django.contrib import messages

from django.contrib.auth.models import User
from django.contrib import messages
from django.shortcuts import render, redirect
from .forms import TrainerForm
from .models import Account

from django.contrib.auth.hashers import make_password


from django.contrib.auth.models import User
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.hashers import make_password
from .forms import TrainerForm
from .models import Account

from django.contrib.auth.models import User
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.hashers import make_password
from .forms import TrainerForm
from .models import Account

from django.contrib.auth.models import User
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.hashers import make_password
from .forms import TrainerForm
from .models import Account

def create_trainer(request):
    if request.method == 'POST':
        form = TrainerForm(request.POST)
        if form.is_valid():
            # Извлекаем значения полей формы
            first_name = form.cleaned_data.get('first_name')
            surname = form.cleaned_data.get('surname')
            patronymic = form.cleaned_data.get('patronymic')
            date_of_birth = form.cleaned_data.get('date_of_birth')
            password = form.cleaned_data.get('password')
            email = form.cleaned_data.get('email')

            username = f"{first_name}"

            # Проверка на существование пользователя
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Пользователь с таким именем уже существует.')
            elif User.objects.filter(email=email).exists():
                messages.error(request, 'Пользователь с таким email уже существует.')
            else:
                # Создаем нового пользователя
                user = User(
                    username=username,
                    email=email,
                    first_name=first_name,
                    password=make_password(password)  # Хешируем пароль
                )
                user.save()  # Сохраняем пользователя

                # Создаем и сохраняем аккаунт с привязкой к пользователю
                account = Account(
                    user=user,
                    surname=surname,
                    patronymic=patronymic,
                    date_of_birth=date_of_birth
                )
                account.save()  # Сохраняем аккаунт

                messages.success(request, 'Тренер успешно создан!')
                return redirect('profile')  # Перенаправляем на страницу профиля
        else:
            messages.error(request, 'Форма содержит ошибки.')
    else:
        form = TrainerForm()

    return render(request, 'SportViolations/create_trainer.html', {'form': form})




def trainer_list(request):
    trainers = Account.objects.filter(role='trainer')  # Получаем только тренеров

    return render(request, 'SportViolations/trainer_list.html', {'trainers': trainers})
